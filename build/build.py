"""
build.py
--------
Builds index.html, sitemap.xml, robots.txt, llms.txt, and Markdown mirrors from:
  - build/template.html        (Jinja2 HTML template)
  - content.yaml               (editable site content)
  - Dropbox XLSX (live)        (race results + calendar)
  - images/gallery/            (photo files)

Usage:
  python3 build/build.py
"""

import sys
import io
import re
import html
import json
import datetime
import requests
from pathlib import Path

import yaml
from jinja2 import Environment, FileSystemLoader
from PIL import Image
from openpyxl.cell.rich_text import CellRichText, TextBlock

# ─── Paths ───────────────────────────────────────────────────────────────────
ROOT = Path(__file__).parent.parent
BUILD_DIR = ROOT / "build"
IMAGES_DIR = ROOT / "images"
GALLERY_DIR = IMAGES_DIR / "gallery"
MARKDOWN_MIRROR_FILES = [
    "index.md",
    "profile.md",
    "results.md",
    "calendar.md",
    "media-kit.md",
    "partnerships.md",
    "gallery.md",
    "values.md",
    "mission.md",
]

# ─── Dropbox Excel config ─────────────────────────────────────────────────────
DROPBOX_XLSX_URL = (
    "https://www.dropbox.com/scl/fi/eaob1d71j6254uz0oocn8/Racing-Calendar.xlsx"
    "?rlkey=zczjnw5wnongwf6091jj82lpp&st=50slpji0&dl=1"
)
SHEET_NAMES = ["2025", "2026", "2027"]

# The year tab active by default on the race results section
CURRENT_YEAR = "2026"

# Alex's UTMB runner profile URL — used to auto-fetch the live UTMB Index at build time
UTMB_RUNNER_URL = "https://utmb.world/en/runner/8058767.alex.schubach"

# Alex's ITRA runner profile URL — used to auto-fetch the live ITRA Performance Index at build time
ITRA_RUNNER_URL = "https://itra.run/RunnerSpace/schubach.alex.6900188"


def _fmt_cell(v) -> str:
    """Convert a cell value to a clean string; formats dates as '1 Jan 2025', times as H:MM:SS."""
    if v is None:
        return ""
    if isinstance(v, datetime.time):
        return v.strftime("%-H:%M:%S").strip()
    if isinstance(v, datetime.timedelta):
        total = int(v.total_seconds())
        h, rem = divmod(total, 3600)
        m, s = divmod(rem, 60)
        return f"{h}:{m:02d}:{s:02d}"
    if hasattr(v, "strftime"):  # datetime.datetime or datetime.date
        return v.strftime("%-d %b %Y").strip()
    return str(v).strip()


# Columns whose native Excel bold formatting should render as <strong> in HTML
RICH_TEXT_COLS = {
    "COMMENTS PRE", "COMMENTS POST", "GOING IN", "LOOKING BACK",
    "PRE RACE", "POST RACE",
    "RACE DESCRIPTION", "DESCRIPTION",
}


def _fmt_narrative(text: str) -> str:
    """Convert **bold** markdown and newlines to HTML for narrative fields."""
    if not text:
        return text
    text = re.sub(r'\*\*(.*?)\*\*', r'<strong>\1</strong>', text)
    text = text.replace('\n', '<br>')
    return text


def _plain_text(value: str) -> str:
    """Convert simple site HTML fragments into readable plain text/Markdown."""
    if not value:
        return ""
    text = str(value)
    text = re.sub(r'<\s*br\s*/?\s*>', '\n', text, flags=re.IGNORECASE)
    text = re.sub(r'</p\s*>', '\n', text, flags=re.IGNORECASE)
    text = re.sub(r'<[^>]+>', '', text)
    text = html.unescape(text)
    text = re.sub(r'[ \t]+\n', '\n', text)
    text = re.sub(r'\n{3,}', '\n\n', text)
    return text.strip()


def _md_table_escape(value: str) -> str:
    return _plain_text(value).replace("|", "\\|").replace("\n", " ")


def _url(base_url: str, path: str = "") -> str:
    base = base_url.rstrip("/")
    return f"{base}/{path.lstrip('/')}" if path else f"{base}/"


def _fmt_rich_cell(cell) -> str:
    """Convert a Cell to HTML, preserving native Excel bold as <strong>.

    openpyxl returns a CellRichText object when the cell contains inline
    formatting (e.g. some words bolded). Elements in that object can be
    either TextBlock (has .font) or plain str (no .font). We must guard
    with isinstance before accessing .font.
    """
    v = cell.value
    if v is None:
        return ""
    if isinstance(v, CellRichText):
        parts = []
        for block in v:
            if isinstance(block, TextBlock):
                text = str(block.text)
                safe = html.escape(text)
                if block.font and getattr(block.font, 'b', False):
                    parts.append(f"<strong>{safe}</strong>")
                else:
                    parts.append(safe)
            else:
                # Bare str element inside CellRichText — no .font attribute
                parts.append(html.escape(str(block)))
        return "".join(parts).replace('\n', '<br>')
    # Plain string fallback: escape HTML special chars first so literal '<'/'>'/'&'
    # in the cell are neutralised before _fmt_narrative applies bold and \n→<br>.
    return _fmt_narrative(html.escape(str(v))) if v else ""


def fetch_excel_sheets(url: str) -> dict:
    """Download XLSX from Dropbox and return {sheet_name: [row_dicts]}."""
    import openpyxl
    try:
        resp = requests.get(url, timeout=30)
        resp.raise_for_status()
        wb = openpyxl.load_workbook(io.BytesIO(resp.content), data_only=True, rich_text=True)
        result = {}
        for name in SHEET_NAMES:
            if name not in wb.sheetnames:
                continue
            ws = wb[name]
            rows = list(ws.iter_rows())
            if not rows:
                result[name] = []
                continue
            # Find header row (first non-empty row)
            header_idx = next((i for i, r in enumerate(rows) if any(c.value for c in r)), None)
            if header_idx is None:
                result[name] = []
                continue
            headers = [str(c.value).strip() if c.value else "" for c in rows[header_idx]]
            result[name] = [
                {
                    headers[i]: (
                        _fmt_rich_cell(cell)
                        if headers[i].upper() in RICH_TEXT_COLS
                        else _fmt_cell(cell.value)
                    )
                    for i, cell in enumerate(row)
                    if i < len(headers)
                }
                for row in rows[header_idx + 1:]
                if any(cell.value for cell in row)
            ]
        return result
    except Exception as e:
        print(f"  ⚠ Could not fetch Excel from Dropbox: {e}", file=sys.stderr)
        return {}


def fetch_utmb_index(runner_url: str, fallback: str = "") -> str:
    """Fetch live UTMB Index from runner's utmb.world profile via __NEXT_DATA__ JSON."""
    import json
    try:
        headers = {
            "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
                          "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
        }
        resp = requests.get(runner_url, timeout=15, headers=headers)
        resp.raise_for_status()
        m = re.search(r'<script id="__NEXT_DATA__"[^>]*>(.*?)</script>', resp.text, re.DOTALL)
        if not m:
            print("  ⚠ UTMB: __NEXT_DATA__ not found in page", file=sys.stderr)
            return fallback
        data = json.loads(m.group(1))
        perf = data["props"]["pageProps"]["performanceIndexes"]
        general = next((entry for entry in perf if entry.get("piCategory") == "general"), None)
        if general and isinstance(general.get("index"), int):
            best = str(general["index"])
            print(f"  UTMB Index fetched: {best} (raw: {perf})")
            return best
    except Exception as e:
        print(f"  ⚠ Could not fetch UTMB index: {e}", file=sys.stderr)
    return fallback


def fetch_itra_index(runner_url: str, fallback_score: str = "", fallback_level: str = "") -> tuple[str, str]:
    """Fetch live ITRA Performance Index and level from runner's itra.run profile.
    Returns (score, level) e.g. ("609", "Advanced 3"). Falls back to provided values on error."""
    try:
        headers = {
            "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
                          "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
        }
        resp = requests.get(runner_url, timeout=15, headers=headers)
        resp.raise_for_status()
        score_match = re.search(r'"performanceIndex":(\d+)', resp.text)
        level_match = re.search(r'"(Advanced|Elite|Expert|Finisher|Recreational)-(\d+)"', resp.text)
        score = score_match.group(1) if score_match else fallback_score
        level = f"{level_match.group(1)} {level_match.group(2)}" if level_match else fallback_level
        if score:
            print(f"  ITRA Index fetched: {score} ({level})")
        return score, level
    except Exception as e:
        print(f"  ⚠ Could not fetch ITRA index: {e}", file=sys.stderr)
    return fallback_score, fallback_level


def optimize_images():
    """Resize and compress all gallery images to max 1200px wide, quality 82.
    PNGs are converted to JPEG (far better compression for photos).
    Uppercase extensions (.JPG, .PNG) are normalised to lowercase .jpg.
    Modifies in-place; originals are replaced/deleted after conversion.
    """
    all_images = (
        list(GALLERY_DIR.glob("*.jpg")) + list(GALLERY_DIR.glob("*.JPG")) +
        list(GALLERY_DIR.glob("*.jpeg")) + list(GALLERY_DIR.glob("*.JPEG")) +
        list(GALLERY_DIR.glob("*.png")) + list(GALLERY_DIR.glob("*.PNG"))
    )

    # Deduplicate by lowercase stem (macOS case-insensitive FS returns .JPG from *.jpg glob)
    seen = set()
    deduped = []
    for p in all_images:
        key = p.stem.lower()
        if key not in seen:
            seen.add(key)
            deduped.append(p)

    for img_path in deduped:
        try:
            img = Image.open(img_path)
            img.load()  # force full read before closing the file handle

            is_png = img_path.suffix.lower() == ".png"
            out_path = img_path.parent / (img_path.stem + ".jpg")

            # Flatten transparency (PNGs may have alpha channel)
            if img.mode in ("RGBA", "LA", "P"):
                if img.mode == "P":
                    img = img.convert("RGBA")
                bg = Image.new("RGB", img.size, (10, 10, 10))
                bg.paste(img, mask=img.split()[-1])
                img = bg
            elif img.mode != "RGB":
                img = img.convert("RGB")

            # Resize if wider than 1200px
            if img.width > 1200:
                ratio = 1200 / img.width
                new_size = (1200, int(img.height * ratio))
                img = img.resize(new_size, Image.LANCZOS)

            if is_png:
                # PNG → JPEG: save to new path, delete original
                img.save(str(out_path), "JPEG", quality=82, optimize=True)
                img_path.unlink()
                print(f"  Converted {img_path.name} → {out_path.name} "
                      f"({out_path.stat().st_size // 1024} KB)")
            else:
                # JPEG (any case extension): save in-place using original path string
                img.save(str(img_path), "JPEG", quality=82, optimize=True)
                print(f"  Optimised {img_path.name} "
                      f"({img_path.stat().st_size // 1024} KB)")

        except Exception as e:
            print(f"  ⚠ Could not optimise {img_path.name}: {e}", file=sys.stderr)

    # Also optimise identity.jpg and about.jpg
    for extra in [IMAGES_DIR / "identity.jpg", IMAGES_DIR / "about.jpg"]:
        if extra.exists():
            try:
                img = Image.open(extra)
                if img.mode != "RGB":
                    img = img.convert("RGB")
                if img.width > 1200:
                    ratio = 1200 / img.width
                    img = img.resize((1200, int(img.height * ratio)), Image.LANCZOS)
                    img.save(extra, "JPEG", quality=82, optimize=True)
            except Exception as e:
                print(f"  ⚠ Could not optimise {extra.name}: {e}", file=sys.stderr)


def _natural_key(p):
    """Sort key for gallery filenames: numeric prefix sorts as integer."""
    m = re.match(r'^(\d+)', p.name)
    return int(m.group(1)) if m else float('inf')


def _build_caption(stem: str) -> str:
    """Parse filename stem into display caption HTML.
    Filmmaker pattern: 'N_Name - Filmmaker_Shot' → two-line HTML (Shot / Photographer).
    Simple pattern: 'N_Description' → title-cased single line.
    """
    name = re.sub(r'^\d+[_\s]*', '', stem)
    if ' - Filmmaker_' in name:
        photographer, shot = name.split(' - Filmmaker_', 1)
        shot_display = shot.replace('_', ' ').title()
        return f"{shot_display}<br>{photographer.strip()}"
    else:
        return name.replace('_', ' ').title()


def build_gallery_meta() -> list[dict]:
    """Scan images/gallery/ and return image metadata for HTML and Markdown outputs."""
    images = sorted(
        list(GALLERY_DIR.glob("*.jpg")) + list(GALLERY_DIR.glob("*.JPG")) +
        list(GALLERY_DIR.glob("*.jpeg")) + list(GALLERY_DIR.glob("*.JPEG")) +
        list(GALLERY_DIR.glob("*.png")) + list(GALLERY_DIR.glob("*.PNG")),
        key=_natural_key
    )
    if not images:
        return []

    meta = []
    for img_path in images:
        try:
            with Image.open(img_path) as im:
                w, h = im.size
        except Exception:
            w, h = 800, 600
        alt = img_path.stem.lstrip("0123456789").strip("_- ").replace("_", " ").title()
        meta.append({"path": img_path, "alt": alt, "caption": _build_caption(img_path.stem), "w": w, "h": h})
    return meta


def build_gallery_html(meta: list[dict] | None = None) -> str:
    """Build featured + thumbnail rail HTML from gallery metadata."""
    if meta is None:
        meta = build_gallery_meta()
    if not meta:
        return "<!-- No gallery images found -->"

    first = meta[0]
    # Featured hero
    featured = (
        f'    <div class="gallery-featured-outer">\n'
        f'      <div class="gallery-featured-main" id="gallery-featured-main">\n'
        f'        <img class="gallery-featured-img" id="gallery-featured-img"'
        f' src="images/gallery/{first["path"].name}" alt="{first["alt"]}">\n'
        f'        <div class="gallery-caption" id="gallery-caption">{first["caption"]}</div>\n'
        f'      </div>\n'
    )

    # Thumbnail rail
    thumbs = ['      <div class="gallery-rail">']
    for idx, m in enumerate(meta):
        active = ' active' if idx == 0 else ''
        thumbs.append(
            f'        <div class="gallery-thumb{active}" data-index="{idx}"'
            f' data-src="images/gallery/{m["path"].name}"'
            f' data-caption="{m["alt"]}" data-caption-html="{html.escape(m["caption"])}"'
            f' data-w="{m["w"]}" data-h="{m["h"]}">'
            f'<img src="images/gallery/{m["path"].name}" alt="{m["alt"]}" loading="lazy"></div>'
        )
    thumbs.append('      </div>')

    # Hidden items list for lightbox (preserves existing lightbox JS query)
    hidden = ['      <div style="display:none" aria-hidden="true">']
    for idx, m in enumerate(meta):
        hidden.append(
            f'        <div class="gallery-item" data-index="{idx}" data-src="images/gallery/{m["path"].name}">'
            f'<img src="" alt="{m["alt"]}"></div>'
        )
    hidden.append('      </div>')

    return featured + "\n".join(thumbs) + "\n" + "\n".join(hidden) + "\n    </div>"


# ─── Excel/Sheets column name helpers ────────────────────────────────────────

def find_col(row: dict, *candidates) -> str:
    """Find first matching column key (case-insensitive, strips whitespace)."""
    keys_lower = {k.strip().lower(): k for k in row.keys() if k is not None}
    for candidate in candidates:
        if candidate.lower() in keys_lower:
            val = row[keys_lower[candidate.lower()]]
            return (val or "").strip()
    return ""


def parse_race_rows(rows: list[dict]) -> tuple[list[dict], list[dict]]:
    """
    Split sheet rows into (past_races, upcoming_races).
    A row is a past race if RACE RESULTS is non-empty.
    A row is upcoming if RACE RESULTS is empty but EVENT is non-empty.
    Filters to rows where REGISTERED contains 'Alex' OR is empty (TBC upcoming).
    """
    past, upcoming = [], []
    for row in rows:
        event = find_col(row, "EVENT", "Event", "RACE", "Race")
        if not event or event.lower().startswith("event"):  # skip header rows
            continue
        registered = find_col(row, "REGISTERED", "Registered", "REGISTRATION")
        reg_lower = registered.lower()
        # Include if not registered column (older sheets), or "yes"/"tbc"/"alex" in value
        if registered and "yes" not in reg_lower and "tbc" not in reg_lower and "alex" not in reg_lower:
            continue

        result = find_col(row, "RACE RESULTS", "Race Results", "RESULT", "Results", "TIME")
        date_str = find_col(row, "RACE DATE", "BLACKOUT DATES", "DATE", "Date")
        race_type = find_col(row, "RACE TYPE", "TYPE") or infer_race_type(event.split("\n")[0].strip())
        col_distance = find_col(row, "RACE DISTANCE", "DISTANCE")
        description = find_col(row, "RACE DESCRIPTION", "DESCRIPTION")
        location = find_col(row, "RACE LOCATION", "LOCATION", "VENUE", "CITY")
        comments_pre = find_col(row, "COMMENTS PRE", "Comments Pre", "PRE RACE", "GOING IN")
        comments_post = find_col(row, "COMMENTS POST", "Comments Post", "POST RACE", "LOOKING BACK")
        pos_overall = find_col(row, "POSITION OVERALL", "Position Overall", "OVERALL POSITION", "POS OVERALL")
        pos_ag = find_col(row, "POSITION AG", "Position AG", "AG POSITION", "AGE GROUP POS")

        race_name = event.split("\n")[0].strip()
        race_date = date_str.split("\n")[0].strip() if date_str else ""
        distance = col_distance or infer_distance(race_name)

        entry = {
            "name": race_name,
            "date": race_date,
            "type": race_type,
            "result": result,
            "description": description,
            "registered": registered,
            "location": location,
            "comments_pre": comments_pre,
            "comments_post": comments_post,
            "pos_overall": pos_overall,
            "pos_ag": pos_ag,
            "distance": distance,
        }

        if result:
            past.append(entry)
        elif race_name:
            upcoming.append(entry)

    return past, upcoming


def infer_race_type(name: str) -> str:
    """Guess race type from name (fallback when RACE TYPE column is absent)."""
    n = name.lower()
    if "hyrox" in n:
        return "Hybrid"
    if "spartan" in n:
        return "OCR"
    if "marathon" in n and "half" not in n:
        return "Road"
    if "half marathon" in n or "half" in n:
        return "Road"
    if "10k" in n or "10km" in n:
        return "Road"
    if "5k" in n or "5km" in n:
        return "Road"
    if "ultra" in n or "80k" in n or "100k" in n:
        return "Ultra"
    if "trail" in n or "mountain" in n or "fuji" in n or "nikko" in n:
        return "Trail"
    return "Road"


def infer_distance(name: str) -> str:
    """Extract distance label from race name."""
    n = name.lower()
    if "marathon" in n and "half" not in n:
        return "42.2 km"
    if "half marathon" in n or "half" in n:
        return "21.1 km"
    if "80k" in n or "80km" in n:
        return "80 km"
    if "50k" in n or "50km" in n:
        return "50 km"
    if "35k" in n or "35km" in n:
        return "35 km"
    if "30k" in n or "30km" in n:
        return "30 km"
    if "27k" in n or "27km" in n:
        return "27 km"
    if "25k" in n or "25km" in n:
        return "25 km"
    if "21k" in n or "21km" in n:
        return "21.1 km"
    if "20k" in n or "20km" in n:
        return "20 km"
    if "12k" in n or "12km" in n:
        return "12 km"
    if "10k" in n or "10km" in n:
        return "10 km"
    if "5k" in n or "5km" in n:
        return "5 km"
    return "—"


def build_race_card_html(race: dict) -> str:
    """Build a single collapsible race card."""
    name = race["name"]
    date = race["date"]
    result = race["result"]
    race_type = race["type"]
    distance = race.get("distance", "—")
    pos = race["pos_overall"] or "—"
    dist_already_in_type = distance.lower().replace(" ", "").replace(".", "") in race_type.lower().replace(" ", "").replace(".", "")
    type_display = f"{race_type}\n{distance}" if distance and distance not in ("—", "") and not dist_already_in_type else race_type
    # HTML-escape header display fields (narrative fields are already escaped by _fmt_rich_cell)
    name = html.escape(name)
    date = html.escape(date)
    result = html.escape(result)
    pos = html.escape(pos)
    type_display = html.escape(type_display)

    # Narrative body
    body_parts = []
    if race.get("description"):
        # Value is already escaped + bold-formatted by _fmt_rich_cell at load time.
        body_parts.append(
            f'              <div class="race-narrative">\n'
            f'                <div class="race-narrative-label">Race Description</div>\n'
            f'                <p>{race["description"]}</p>\n'
            f'              </div>'
        )
    if race["comments_pre"]:
        body_parts.append(
            f'              <div class="race-narrative">\n'
            f'                <div class="race-narrative-label">Going In</div>\n'
            f'                <p>{race["comments_pre"]}</p>\n'
            f'              </div>'
        )
    if race["comments_post"]:
        body_parts.append(
            f'              <div class="race-narrative">\n'
            f'                <div class="race-narrative-label">Looking Back</div>\n'
            f'                <p>{race["comments_post"]}</p>\n'
            f'              </div>'
        )

    body_html = "\n".join(body_parts) if body_parts else "              <p>Race notes coming soon.</p>"

    location = html.escape(race.get("location", "") or "")
    loc_html = f'<div class="race-h-loc">{location}</div>' if location and location != "—" else ""
    has_desc_class = (" has-desc" if race.get("description") else "")

    return f'''        <div class="race-item">
          <div class="race-header" tabindex="0" role="button" aria-expanded="false">
            <div><div class="race-h-name">{name}</div><div class="race-h-sub">{date}</div>{loc_html}</div>
            <div class="race-h-type">{type_display}</div>
            <div class="race-h-time">{result or "—"}</div>
            <div class="race-h-pos">{pos}</div>
            <div class="race-h-expand">Expand</div>
          </div>
          <div class="race-body">
            <div class="race-body-inner{has_desc_class}">
{body_html}
            </div>
          </div>
        </div>'''


def build_calendar_card_html(race: dict) -> str:
    """Build a calendar card for an upcoming race."""
    tbc = "tbc" in race.get("registered", "").lower()
    tbc_badge = '<span class="cal-tbc">TBC</span>' if tbc else ""
    detail_parts = [p for p in [race.get("type"), race.get("distance"), race.get("location")] if p and p != "—"]
    details = " · ".join(detail_parts)

    desc = race.get("description", "")
    if desc:
        # Strip HTML tags and unescape entities to get clean plain text for the
        # calendar preview and modal data-desc attribute. The modal uses
        # textContent (not innerHTML) so we need literal characters, not entities.
        desc_plain = html.unescape(re.sub(r'<[^>]+>', '', desc))
        desc_html = f'        <div class="cal-desc">{html.escape(desc_plain)}</div>\n'
        data_attrs = (
            f' data-desc="{html.escape(desc_plain)}"'
            f' data-race="{html.escape(race["name"].upper())}"'
            f' data-date="{html.escape(race["date"])}"'
            f' data-details="{html.escape(details)}"'
        )
        read_more_html = '        <button class="cal-read-more" onclick="openCalModal(this)">Read more →</button>\n'
    else:
        desc_html = ""
        data_attrs = ""
        read_more_html = ""

    return (
        f'      <div class="cal-card reveal"{data_attrs}>\n'
        f'        <div class="cal-month">{race["date"]} {tbc_badge}</div>\n'
        f'        <div class="cal-race">{race["name"].upper()}</div>\n'
        f'        <div class="cal-details">{details}</div>\n'
        f'{desc_html}'
        f'{read_more_html}'
        f'      </div>'
    )


TABLE_HEADER = '''      <div class="race-table-header">
        <div class="race-th">Race</div>
        <div class="race-th">Type</div>
        <div class="race-th">Time</div>
        <div class="race-th">Ranking</div>
        <div class="race-th race-th-expand">Expand</div>
      </div>'''


def build_race_tabs_and_panels(sheets_data: dict) -> tuple:
    """
    Build the year-tabs + year-panels HTML from all sheets.
    sheets_data = { "2025": (past_races, upcoming), "2026": (past_races, upcoming), ... }
    Returns (tabs_and_panels_html, calendar_cards_html, calendar_year)
    """
    # Sort years ascending so tabs appear 2025 → 2026 → ...
    years = sorted(sheets_data.keys())

    tabs_html = '<div class="year-tabs reveal">\n'
    panels_html = ""
    all_upcoming = []

    for sheet_name in years:
        year_full = sheet_name  # already "2025", "2026", etc.
        past_races, upcoming_races = sheets_data[sheet_name]
        is_current = year_full == CURRENT_YEAR

        tabs_html += f'  <div class="year-tab{" active" if is_current else ""}" data-year="{year_full}">{year_full}</div>\n'

        race_cards = "\n".join(build_race_card_html(r) for r in reversed(past_races)) if past_races else \
            '        <p style="color:var(--grey-mid);padding:2rem 0">No results recorded yet.</p>'

        panels_html += (
            f'\n    <div class="year-panel{" active" if is_current else ""}" id="panel-{year_full}">\n'
            f'{TABLE_HEADER}\n'
            f'      <div class="race-accordion">\n'
            f'{race_cards}\n'
            f'      </div>\n'
            f'    </div>'
        )

        all_upcoming.extend(upcoming_races)

    tabs_html += "</div>"

    calendar_html = "\n".join(build_calendar_card_html(r) for r in all_upcoming) if all_upcoming else \
        '      <p style="color:var(--grey-mid)">Calendar coming soon.</p>'

    return tabs_html + panels_html, calendar_html, CURRENT_YEAR


def build_seo_tags(site: dict, social: dict, analytics: dict = None) -> str:
    """Build Open Graph, Twitter Card, JSON-LD, and GSC verification tags."""
    base_url = site.get("base_url", "")
    title = site.get("title", "Alex Schubach — Endurance Athlete")
    description = site.get("description", "")
    instagram = social.get("instagram", "#")
    strava = social.get("strava", "#")

    gsc_token = (analytics or {}).get("gsc_verification_token", "").strip()
    gsc_tag = f'  <meta name="google-site-verification" content="{gsc_token}">\n' if gsc_token else ""

    title_esc = html.escape(title)
    desc_esc = html.escape(description)

    og = gsc_tag + f'''  <!-- Keywords -->
  <meta name="keywords" content="Alex Schubach, Alexander Schubach, Alex the Athlete, alexschubach, alexschubach.com, Tokyo endurance athlete, performance athlete, athlete model, athlete modelling, sponsorship profile, brand partner, endurance athlete, trail runner, Hyrox, ultra running, UTMB, Spartan">
  <!-- Open Graph -->
  <meta property="og:title" content="{title_esc}">
  <meta property="og:description" content="{desc_esc}">
  <meta property="og:image" content="{base_url}/images/hero.jpg">
  <meta property="og:url" content="{base_url}">
  <meta property="og:type" content="website">
  <meta name="twitter:card" content="summary_large_image">
  <meta name="twitter:title" content="{title_esc}">
  <meta name="twitter:description" content="{desc_esc}">
  <meta name="twitter:image" content="{base_url}/images/hero.jpg">'''

    same_as = [s for s in [instagram, strava] if s and s != "#"]
    jsonld_data = {
        "@context": "https://schema.org",
        "@type": "Person",
        "name": "Alex Schubach",
        "alternateName": ["Alexander Schubach", "Alex the Athlete", "alexschubach"],
        "url": base_url,
        "image": f"{base_url}/images/about.jpg",
        "jobTitle": "Endurance athlete, performance athlete, athlete model, and brand partner",
        "description": description,
        "homeLocation": {
            "@type": "Place",
            "name": "Tokyo, Japan",
        },
        "workLocation": {
            "@type": "Place",
            "name": "Japan and Asia-Pacific",
        },
        "hasOccupation": [
            {"@type": "Occupation", "name": "Endurance Athlete"},
            {"@type": "Occupation", "name": "Performance Athlete"},
            {"@type": "Occupation", "name": "Athlete Model"},
            {"@type": "Occupation", "name": "Brand Partner"},
        ],
        "knowsAbout": [
            "endurance sport",
            "trail running",
            "road running",
            "Hyrox",
            "Spartan racing",
            "ultra running",
            "athlete modelling",
            "brand partnerships",
            "sports sponsorship",
            "fitness content",
            "outdoor performance",
            "Japan and Asia-Pacific racing",
        ],
        "sameAs": same_as,
    }
    jsonld = (
        "  <!-- JSON-LD Structured Data -->\n"
        "  <script type=\"application/ld+json\">\n"
        + json.dumps(jsonld_data, indent=2, ensure_ascii=False)
        + "\n  </script>"
    )

    return og + "\n" + jsonld


def build_analytics_tags(analytics: dict) -> str:
    """Build GA4 or Plausible analytics script tag."""
    ga4_id = analytics.get("ga4_id", "").strip()
    plausible_domain = analytics.get("plausible_domain", "").strip()

    if ga4_id:
        return f'''  <!-- Google Analytics 4 -->
  <script async src="https://www.googletagmanager.com/gtag/js?id={ga4_id}"></script>
  <script>
    window.dataLayer = window.dataLayer || [];
    function gtag(){{dataLayer.push(arguments);}}
    gtag('js', new Date());
    gtag('config', '{ga4_id}');
  </script>'''
    elif plausible_domain:
        return f'  <script defer data-domain="{plausible_domain}" src="https://plausible.io/js/script.js"></script>'

    return "  <!-- Analytics: set ga4_id or plausible_domain in content.yaml -->"


def build_llms_txt(base_url: str, content: dict, indices: dict) -> str:
    site = content.get("site", {})
    social = content.get("social", {})
    pdf = content.get("pdf", {})
    contact = content.get("contact", {})
    positioning = content.get("positioning", {})
    description = site.get("description", "")
    itra_label = indices.get("itra") or "not listed"
    if indices.get("itra_level"):
        itra_label = f"{itra_label} ({indices.get('itra_level')})"

    lines = [
        "# Alex Schubach",
        "",
        f"> {_plain_text(description)}",
        "",
        "Alex Schubach is a Tokyo-based performance athlete, endurance athlete, trail runner, Hyrox competitor, athlete model, and brand partner. This file points AI systems to stable, text-first mirrors of the public site.",
        "",
        _plain_text(positioning.get("summary", "")),
        "",
        "## Core Pages",
        "",
        f"- [Homepage]({_url(base_url, 'index.md')}): Athlete identity, headline metrics, race results, calendar, gallery, downloads, and sponsorship/contact routes.",
        f"- [Profile]({_url(base_url, 'profile.md')}): Athletic background, Tokyo location, disciplines, athlete modelling context, and brand partner positioning.",
        f"- [Results]({_url(base_url, 'results.md')}): Performance indices, personal bests, and race results.",
        f"- [Calendar]({_url(base_url, 'calendar.md')}): Upcoming races and target events.",
        f"- [Media Kit]({_url(base_url, 'media-kit.md')}): Downloadable media kit and training/nutrition resources for partnership review.",
        f"- [Partnerships]({_url(base_url, 'partnerships.md')}): Sponsorship enquiries, athlete modelling briefs, media, collaborations, and management contact route.",
        f"- [Gallery]({_url(base_url, 'gallery.md')}): Public image references and visual profile.",
        f"- [Values]({_url(base_url, 'values.md')}): Alex's stated values.",
        f"- [Mission]({_url(base_url, 'mission.md')}): Mission, mindset, and performance standard.",
        "",
        "## Performance Facts",
        "",
        f"- Based in Tokyo, Japan.",
        "- Disciplines: trail running, ultra running, road running, Hyrox, Spartan, obstacle racing, and hybrid endurance.",
        f"- UTMB Index: {indices.get('utmb') or 'not listed'}.",
        f"- ITRA Index: {itra_label}.",
        f"- Hyrox PB: {indices.get('hyrox_pb') or 'not listed'}.",
        f"- Marathon PB: {indices.get('road_marathon') or 'not listed'}.",
        f"- Half marathon PB: {indices.get('road_half') or 'not listed'}.",
        f"- 10 km PB: {indices.get('road_10k') or 'not listed'}.",
        f"- 5 km PB: {indices.get('road_5k') or 'not listed'}.",
        "",
        "## Partnership Fit",
        "",
        "- Sponsorship, ambassador, affiliate, product-testing, event, media, athlete modelling, and campaign enquiries.",
        f"- {_plain_text(positioning.get('brand_fit', 'Relevant categories: endurance sport, trail running, road running, Hyrox, outdoor gear, nutrition, recovery, wearables, travel, Japan/Asia-Pacific races, modelling, and content production.'))}",
        "",
        "## Contact",
        "",
        f"- Website: {_url(base_url)}",
        "- Partnership contact: manager@alexschubach.com",
        f"- Instagram: {social.get('instagram') or 'not listed'}",
        f"- Strava: {social.get('strava') or 'not listed'}",
        f"- Blog: {social.get('blog') or 'not listed'}",
    ]

    media_kit = pdf.get("media_kit_url")
    if media_kit:
        lines.append(f"- Media kit: {_url(base_url, media_kit)}")

    if contact.get("categories"):
        lines.extend(["", "## Enquiry Types", ""])
        for category in contact.get("categories", []):
            lines.append(f"- {category.get('label', 'Enquiry')}: {_plain_text(category.get('desc', ''))}")

    lines.extend([
        "",
        "## Optional",
        "",
        f"- [Sitemap]({_url(base_url, 'sitemap.xml')}): XML sitemap.",
        f"- [Robots]({_url(base_url, 'robots.txt')}): Current crawler policy.",
    ])

    return "\n".join(lines).strip() + "\n"


def build_markdown_mirrors(base_url: str, content: dict, indices: dict, sheets_data: dict, gallery_meta: list[dict]) -> dict:
    site = content.get("site", {})
    about = content.get("about", {})
    values = content.get("values", [])
    mission = content.get("mission", {})
    contact = content.get("contact", {})
    pdf = content.get("pdf", {})
    social = content.get("social", {})
    positioning = content.get("positioning", {})
    today = datetime.date.today().isoformat()
    itra_label = indices.get("itra") or "not listed"
    if indices.get("itra_level"):
        itra_label = f"{itra_label} ({indices.get('itra_level')})"

    all_past = []
    all_upcoming = []
    for year in sorted(sheets_data.keys()):
        past, upcoming = sheets_data[year]
        all_past.extend((year, race) for race in past)
        all_upcoming.extend((year, race) for race in upcoming)

    profile_lines = [
        "# Alex Schubach Profile",
        "",
        f"Last updated: {today}",
        f"Canonical URL: {_url(base_url, '#about')}",
        "",
        _plain_text(site.get("description", "")),
        "",
        _plain_text(positioning.get("summary", "")),
        "",
        f"Hero: {_plain_text(content.get('hero', {}).get('subtitle', ''))}",
        "",
        f"## {about.get('heading', 'Athletic Identity')}",
        "",
        *[_plain_text(p) for p in about.get("paragraphs", [])],
        "",
        "## Stats",
        "",
    ]
    for stat in about.get("stats", []):
        profile_lines.append(f"- {stat.get('label')}: {stat.get('number')}")

    values_lines = [
        "# Alex Schubach Values",
        "",
        f"Last updated: {today}",
        f"Canonical URL: {_url(base_url, '#values')}",
        "",
    ]
    for value in values:
        values_lines.extend([
            f"## {value.get('num', '')}. {value.get('title', '')}".strip(),
            "",
            _plain_text(value.get("desc", "")),
            "",
        ])

    mission_lines = [
        "# Alex Schubach Mission",
        "",
        f"Last updated: {today}",
        f"Canonical URL: {_url(base_url, '#mission')}",
        "",
        _plain_text(mission.get("quote", "")),
        "",
        _plain_text(mission.get("body", "")),
        "",
        f"## {mission.get('pillars_heading', 'The Standard')}",
        "",
    ]
    for pillar in mission.get("pillars", []):
        mission_lines.append(f"- {_plain_text(pillar)}")

    results_lines = [
        "# Alex Schubach Results",
        "",
        f"Last updated: {today}",
        f"Canonical URL: {_url(base_url, '#results')}",
        "",
        "## Performance Athlete Summary",
        "",
        _plain_text(positioning.get("summary", "")),
        "",
        *[f"- {_plain_text(item)}" for item in positioning.get("highlights", [])],
        "",
        "## Performance Indices and Personal Bests",
        "",
        f"- UTMB Index: {indices.get('utmb') or 'not listed'}",
        f"- ITRA Index: {itra_label}",
        f"- Hyrox PB: {indices.get('hyrox_pb') or 'not listed'}",
        f"- Marathon PB: {indices.get('road_marathon') or 'not listed'}",
        f"- Half marathon PB: {indices.get('road_half') or 'not listed'}",
        f"- 10 km PB: {indices.get('road_10k') or 'not listed'}",
        f"- 5 km PB: {indices.get('road_5k') or 'not listed'}",
        "",
        "## Race Results",
        "",
        "| Year | Date | Event | Type | Distance | Result | Overall | Age Group | Location |",
        "| --- | --- | --- | --- | --- | --- | --- | --- | --- |",
    ]
    if all_past:
        for year, race in sorted(all_past, key=lambda item: (item[0], item[1].get("date", "")), reverse=True):
            results_lines.append(
                "| " + " | ".join([
                    _md_table_escape(year),
                    _md_table_escape(race.get("date", "")),
                    _md_table_escape(race.get("name", "")),
                    _md_table_escape(race.get("type", "")),
                    _md_table_escape(race.get("distance", "")),
                    _md_table_escape(race.get("result", "")),
                    _md_table_escape(race.get("pos_overall", "")),
                    _md_table_escape(race.get("pos_ag", "")),
                    _md_table_escape(race.get("location", "")),
                ]) + " |"
            )
    else:
        results_lines.append("| | | Race data temporarily unavailable. | | | | | | |")

    calendar_lines = [
        "# Alex Schubach Race Calendar",
        "",
        f"Last updated: {today}",
        f"Canonical URL: {_url(base_url, '#calendar')}",
        "",
        _plain_text(content.get("calendar", {}).get("intro", "")),
        "",
        "| Year | Date | Event | Type | Distance | Location | Status |",
        "| --- | --- | --- | --- | --- | --- | --- |",
    ]
    if all_upcoming:
        for year, race in sorted(all_upcoming, key=lambda item: (item[0], item[1].get("date", ""))):
            calendar_lines.append(
                "| " + " | ".join([
                    _md_table_escape(year),
                    _md_table_escape(race.get("date", "")),
                    _md_table_escape(race.get("name", "")),
                    _md_table_escape(race.get("type", "")),
                    _md_table_escape(race.get("distance", "")),
                    _md_table_escape(race.get("location", "")),
                    _md_table_escape(race.get("registered", "")),
                ]) + " |"
            )
    else:
        calendar_lines.append("| | | Race calendar temporarily unavailable. | | | | |")

    media_lines = [
        "# Alex Schubach Media Kit and Downloads",
        "",
        f"Last updated: {today}",
        f"Canonical URL: {_url(base_url, '#pdf')}",
        "",
        "Use these assets to review Alex Schubach's Tokyo-based endurance athlete profile for sponsorship, brand partnership, athlete modelling, performance, outdoor, nutrition, travel, and lifestyle opportunities.",
        "",
    ]
    download_labels = {
        "media_kit_url": "Media Kit",
        "training_program_url": "Hybrid Training Split",
        "leg_conditioning_url": "Leg Conditioning Program",
        "meal_plan_url": "Maintenance Meal Plan",
        "supplements_url": "Electrolytes and Supplements",
    }
    for key, label in download_labels.items():
        if pdf.get(key):
            media_lines.append(f"- [{label}]({_url(base_url, pdf[key])})")

    partnerships_lines = [
        "# Partner With Alex Schubach",
        "",
        f"Last updated: {today}",
        f"Canonical URL: {_url(base_url, '#contact')}",
        "",
        "## Partnership Positioning",
        "",
        _plain_text(positioning.get("summary", "")),
        "",
        _plain_text(positioning.get("brand_fit", "")),
        "",
        "## Contact Context",
        "",
        _plain_text(contact.get("body", "")),
        "",
        "## Enquiry Categories",
        "",
    ]
    for category in contact.get("categories", []):
        partnerships_lines.extend([
            f"### {category.get('label', 'Enquiry')}",
            "",
            _plain_text(category.get("desc", "")),
            "",
        ])
    partnerships_lines.extend([
        "## Contact",
        "",
        "- Email: manager@alexschubach.com",
        f"- Instagram: {social.get('instagram') or 'not listed'}",
        f"- Strava: {social.get('strava') or 'not listed'}",
        f"- Blog: {social.get('blog') or 'not listed'}",
    ])

    gallery_lines = [
        "# Alex Schubach Gallery",
        "",
        f"Last updated: {today}",
        f"Canonical URL: {_url(base_url, '#gallery')}",
        "",
        "Public gallery image references from alexschubach.com.",
        "",
    ]
    for item in gallery_meta:
        image_path = f"images/gallery/{item['path'].name}"
        gallery_lines.append(f"- [{_plain_text(item['caption'])}]({_url(base_url, image_path)})")

    index_lines = [
        "# Alex Schubach",
        "",
        f"Last updated: {today}",
        f"Canonical URL: {_url(base_url)}",
        "",
        _plain_text(site.get("description", "")),
        "",
        "## Positioning Summary",
        "",
        _plain_text(positioning.get("summary", "")),
        "",
        _plain_text(positioning.get("brand_fit", "")),
        "",
        "## Site Mirrors",
        "",
        f"- [Profile]({_url(base_url, 'profile.md')})",
        f"- [Values]({_url(base_url, 'values.md')})",
        f"- [Mission]({_url(base_url, 'mission.md')})",
        f"- [Results]({_url(base_url, 'results.md')})",
        f"- [Calendar]({_url(base_url, 'calendar.md')})",
        f"- [Gallery]({_url(base_url, 'gallery.md')})",
        f"- [Media Kit]({_url(base_url, 'media-kit.md')})",
        f"- [Partnerships]({_url(base_url, 'partnerships.md')})",
    ]

    return {
        "index.md": "\n".join(index_lines).strip() + "\n",
        "profile.md": "\n".join(profile_lines).strip() + "\n",
        "values.md": "\n".join(values_lines).strip() + "\n",
        "mission.md": "\n".join(mission_lines).strip() + "\n",
        "results.md": "\n".join(results_lines).strip() + "\n",
        "calendar.md": "\n".join(calendar_lines).strip() + "\n",
        "media-kit.md": "\n".join(media_lines).strip() + "\n",
        "partnerships.md": "\n".join(partnerships_lines).strip() + "\n",
        "gallery.md": "\n".join(gallery_lines).strip() + "\n",
    }


def build_sitemap(base_url: str) -> str:
    today = datetime.date.today().isoformat()
    # Ensure trailing slash removed for consistency
    url = base_url.rstrip("/")
    paths = [""] + ["llms.txt"] + MARKDOWN_MIRROR_FILES
    entries = []
    for path in paths:
        loc = f"{url}/" if not path else f"{url}/{path}"
        priority = "1.0" if not path else ("0.8" if path == "llms.txt" else "0.7")
        entries.append(f'''  <url>
    <loc>{loc}</loc>
    <lastmod>{today}</lastmod>
    <changefreq>weekly</changefreq>
    <priority>{priority}</priority>
  </url>''')
    return f'''<?xml version="1.0" encoding="UTF-8"?>
<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">
{chr(10).join(entries)}
</urlset>
'''


def build_robots(base_url: str) -> str:
    url = base_url.rstrip("/")
    return f'''User-agent: *
Allow: /
Sitemap: {url}/sitemap.xml
'''


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    print("\n=== Alex Schubach site builder ===\n")

    # 1. Load content.yaml
    content_path = ROOT / "content.yaml"
    if not content_path.exists():
        print("Error: content.yaml not found", file=sys.stderr)
        sys.exit(1)
    with open(content_path) as f:
        content = yaml.safe_load(f)

    site = content.get("site", {})
    base_url = site.get("base_url", "").rstrip("/")

    # 1b. Fetch live UTMB index (overrides content.yaml fallback)
    indices = content.get("indices", {})
    print("Fetching UTMB index...")
    live_utmb = fetch_utmb_index(UTMB_RUNNER_URL, fallback=indices.get("utmb", ""))
    if live_utmb:
        indices = {**indices, "utmb": live_utmb}

    # 1c. Fetch live ITRA index (overrides content.yaml fallback)
    print("Fetching ITRA index...")
    live_itra, live_itra_level = fetch_itra_index(
        ITRA_RUNNER_URL,
        fallback_score=indices.get("itra", ""),
        fallback_level=indices.get("itra_level", ""),
    )
    if live_itra:
        indices = {**indices, "itra": live_itra, "itra_level": live_itra_level}

    # 2. Optimise images
    print("Optimising images...")
    optimize_images()

    # 3. Build gallery HTML
    print("Building gallery...")
    gallery_meta = build_gallery_meta()
    gallery_html = build_gallery_html(gallery_meta)

    # 4. Fetch race data from Dropbox Excel
    print("Fetching race data from Dropbox Excel...")
    raw_sheets = fetch_excel_sheets(DROPBOX_XLSX_URL)
    sheets_data = {}
    for sheet_name, rows in raw_sheets.items():
        past, upcoming = parse_race_rows(rows)
        sheets_data[sheet_name] = (past, upcoming)
        print(f"  '{sheet_name}' → {len(past)} past, {len(upcoming)} upcoming")

    # 5. Build race tabs/panels and calendar
    if sheets_data:
        race_tabs_and_panels, calendar_cards, calendar_year = build_race_tabs_and_panels(sheets_data)
    else:
        # Fallback placeholder when Dropbox fetch fails
        race_tabs_and_panels = (
            f'<div class="year-tabs reveal"><div class="year-tab active" data-year="{CURRENT_YEAR}">{CURRENT_YEAR}</div></div>\n'
            f'<div class="year-panel active" id="panel-{CURRENT_YEAR}">'
            '<p style="color:var(--grey-mid);padding:2rem 0">Race data temporarily unavailable — check back soon.</p>'
            '</div>'
        )
        calendar_cards = '<p style="color:var(--grey-mid)">Race calendar temporarily unavailable — check back soon.</p>'
        calendar_year = CURRENT_YEAR

    # 6. Build SEO and analytics tags
    seo_tags = build_seo_tags(site, content.get("social", {}), analytics=content.get("analytics", {}))
    analytics_tags = build_analytics_tags(content.get("analytics", {}))

    # 7. Render template
    print("Rendering template...")
    env = Environment(loader=FileSystemLoader(str(BUILD_DIR)), autoescape=False)
    tmpl = env.get_template("template.html")

    rendered = tmpl.render(
        site=site,
        hero=content.get("hero", {}),
        about=content.get("about", {}),
        values=content.get("values", []),
        mission=content.get("mission", {}),
        indices=indices,
        contact=content.get("contact", {}),
        social=content.get("social", {}),
        pdf=content.get("pdf", {"enabled": False}),
        footer=content.get("footer", {}),
        seo_tags=seo_tags,
        analytics_tags=analytics_tags,
        race_tabs_and_panels=race_tabs_and_panels,
        calendar=content.get("calendar", {}),
        calendar_cards=calendar_cards,
        calendar_year=calendar_year,
        gallery_items=gallery_html,
    )

    # 8. Write output files
    out_html = ROOT / "index.html"
    out_html.write_text(rendered, encoding="utf-8")
    print(f"  ✓ index.html written ({len(rendered) // 1024} KB)")

    out_sitemap = ROOT / "sitemap.xml"
    out_sitemap.write_text(build_sitemap(base_url), encoding="utf-8")
    print("  ✓ sitemap.xml written")

    out_robots = ROOT / "robots.txt"
    out_robots.write_text(build_robots(base_url), encoding="utf-8")
    print("  ✓ robots.txt written")

    out_llms = ROOT / "llms.txt"
    out_llms.write_text(build_llms_txt(base_url, content, indices), encoding="utf-8")
    print("  ✓ llms.txt written")

    for filename, markdown in build_markdown_mirrors(base_url, content, indices, sheets_data, gallery_meta).items():
        (ROOT / filename).write_text(markdown, encoding="utf-8")
    print(f"  ✓ {len(MARKDOWN_MIRROR_FILES)} Markdown mirrors written")

    (ROOT / ".nojekyll").write_text("", encoding="utf-8")
    print("  ✓ .nojekyll written")

    print("\n✓ Build complete.")


if __name__ == "__main__":
    main()
